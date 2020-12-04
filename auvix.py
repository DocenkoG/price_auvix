# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                      # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, nameToId, currencyType, openX, sheetByName
import csv
import requests, lxml.html



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена со скидкой','цена_') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh) == '' :       # .find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        #if item in ('закупка','продажа','цена','цена1') :
        if item.find('цена') >= 0 :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_csv2csv( cfg ):
    inFfileName  = cfg.get('basic', 'filename_in')
    outFfileNameRUR = cfg.get('basic', 'filename_out_RUR')
    outFfileNameEUR = cfg.get('basic', 'filename_out_EUR')
    outFfileNameUSD = cfg.get('basic', 'filename_out_USD')
    #inFile  = open( inFfileName,  'r', newline='', encoding='CP1251', errors='replace')
    inFile  = open( inFfileName,  'r', encoding='UTF-8', errors='replace')
    outFileRUR = open( outFfileNameRUR, 'w', newline='')
    outFileEUR = open( outFfileNameEUR, 'w', newline='')
    outFileUSD = open( outFfileNameUSD, 'w', newline='')
    
    outFields = cfg.options('cols_out')
    csvReader = csv.DictReader(inFile, delimiter=';', fieldnames=[
        'Модель',
        'Артикул',
        'Наименование',
        'Розничная цена, руб.',
        'Дилерская цена, руб.',
        'Гарантия',
        'Наличие(0 - нет / 1 - есть)',
        'Категория',
        'Производитель',
        'Информация'
    ])
    csvWriterRUR = csv.DictWriter(outFileRUR, fieldnames=cfg.options('cols_out'))
    csvWriterEUR = csv.DictWriter(outFileEUR, fieldnames=cfg.options('cols_out'))
    csvWriterUSD = csv.DictWriter(outFileUSD, fieldnames=cfg.options('cols_out'))

    print(csvReader.fieldnames)

    csvWriterRUR.writeheader()
    csvWriterEUR.writeheader()
    csvWriterUSD.writeheader()
    recOut = {}
    for recIn in csvReader:
        for outColName in outFields :
            shablon = cfg.get('cols_out',outColName)
            for key in csvReader.fieldnames:
                if shablon.find(key) >= 0 :
                    shablon = shablon.replace(key, recIn[key])
            if outColName in('закупка','продажа'):
                if shablon.find('Звоните') >=0 :
                    shablon = '0.1'
            recOut[outColName] = shablon
        csvWriterRUR.writerow(recOut)
        #elif recOut['валюта'] == 'EUR' :
        #    csvWriterEUR.writerow(recOut)
        #elif recOut['валюта'] == 'USD' :
        #    csvWriterUSD.writerow(recOut)
        #else :
        #    log.error('нераспознана валюта "%s" для товара "%s"', recOut['валюта'], recOut['код производителя'] )
    log.info('Обработано '+ str(csvReader.line_num) +'строк.')
    inFile.close()
    outFileRUR.close()
    outFileEUR.close()
    outFileUSD.close()


def convert_excel2csv(cfg):
    csvFName = cfg.get('basic', 'filename_out')
    priceFName = cfg.get('basic', 'filename_in')
    sheetName = cfg.get('basic', 'sheetname')

    log.debug('Reading file ' + priceFName)
    sheet = sheetByName(fileName=priceFName, sheetName=sheetName)
    if not sheet:
        log.error("Нет листа " + sheetName + " в файле " + priceFName)
        return False
    log.debug("Sheet   " + sheetName)
    out_cols = cfg.options("cols_out")
    in_cols = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols:
        out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols:
        in_cols_j[vName] = cfg.getint("cols_in", vName)
    # brands,   discount     = config_read(cfgFName, 'discount')
    # for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    # print(discount)

    outFile = open(csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriter = csv.DictWriter(outFile, fieldnames=out_cols)
    csvWriter.writeheader()

    '''                                     # Блок проверки свойств для распознавания групп      XLSX                                  
    for i in range(2393, 2397):                                                         
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['группа'] )
        print(i, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color.rgb, '------', ccc.fill.fgColor.rgb)
        print('------')
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(0, 75):                                                         
        xfx = sheet.cell_xf_index(i, 0)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 0)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 0).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut = {}
    for i in range(1, sheet.max_row + 1):  # xlsx
        #   for i in range(1, sheet.nrows) :                                     # xls
        i_last = i
        try:
            impValues = getXlsxString(sheet, i, in_cols_j)  # xlsx
            #           impValues = getXlsString(sheet, i, in_cols_j)                # xls
            # print( impValues )
            if impValues['цена1'] == '0':  # (ccc.value == None) or (ccc2.value == None) :     # Пустая строка
                pass
                # print( 'Пустая строка. i=',i, impValues )
            elif impValues['код_'] == '' or impValues['код_'] == 'Арт.':  # Пустая строка
                print(i, 'Пусто!!!')
                continue
            else:  # Обычная строка
                for outColName in out_template.keys():
                    shablon = out_template[outColName]
                    for key in impValues.keys():
                        if shablon.find(key) >= 0:
                            shablon = shablon.replace(key, impValues[key])
                    if (outColName == 'закупка') and ('*' in shablon):
                        p = shablon.find("*")
                        vvv1 = float(shablon[:p])
                        vvv2 = float(shablon[p + 1:])
                        shablon = str(round(vvv1 * vvv2, 2))
                    recOut[outColName] = shablon.strip()

                recOut['код'] = nameToId(recOut['код'])
                csvWriter.writerow(recOut)

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) + '.')

    log.info('Обработано ' + str(i_last) + ' строк.')
    outFile.close()



def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def download( cfg ):
    retCode     = False
    filename_new= cfg.get('download','filename_new')
    filename_old= cfg.get('download','filename_old')
    login       = cfg.get('private','login'    )
    password    = cfg.get('private','password' )
    url_lk      = cfg.get('download','url_lk'   )
    url_file    = cfg.get('download','url_file' )
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.0; rv:14.0) Gecko/20100101 Firefox/14.0.1',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
               'Accept-Language':'ru-ru,ru;q=0.8,en-us;q=0.5,en;q=0.3',
               'Accept-Encoding':'gzip, deflate',
               'Connection':'keep-alive',
               'DNT':'1'
              }

    try:
        s = requests.Session()
        r = s.get(url_lk,  headers = headers)  # auth=(login,password),(И без него сработало, но где-то может понадобиться)
        page = lxml.html.fromstring(r.text)
        form = page.forms[0]
        #print(form)
        form.fields['USER_LOGIN'] = login
        form.fields['USER_PASSWORD'] = password
        r = s.post(url_lk+ form.action, data=form.form_values())
        #print('<<<',r.text,'>>>')
        print('       ==================================================')

        log.debug('Авторизация на %s   --- code=%d', url_lk, r.status_code)
        r = s.get(url_file)
        log.debug('Загрузка файла %16d bytes   --- code=%d', len(r.content), r.status_code)
        retCode = True
    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')

    if os.path.exists( filename_new) and os.path.exists( filename_old): 
        os.remove( filename_old)
        os.rename( filename_new, filename_old)
    if os.path.exists( filename_new) :
        os.rename( filename_new, filename_old)
    f2 = open(filename_new, 'wb')                                  # Теперь записываем файл
    f2.write(r.content)
    f2.close()
    if filename_new[-4:] == '.zip':                                # Архив. Обработка не завершена
        log.debug( 'Zip-архив. Разархивируем.')
        #dir_befo_download = set(os.listdir(os.getcwd()))
        os.system('unzip -oj ' + filename_new)
        #dir_afte_download = set(os.listdir(os.getcwd()))
        #new_files = list( dir_afte_download.difference(dir_befo_download))
    return retCode



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    file_age = round((time.time() - price_datetime) / 24 / 60 / 60)
    if file_age > qty_days :
        log.error('Файл "' + fileName + '" устарел! Допустимый период ' + str(qty_days)+' дней, а ему ' + str(file_age))
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def main(dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          ' + dealerName)

    rc_download = False
    '''
    '''
    if os.path.exists('getting.cfg'):
        cfg = config_read('getting.cfg')
        filename_new = cfg.get('basic','filename_new')
        if cfg.has_section('download'):
            rc_download = download(cfg)
        if not(rc_download==True or is_file_fresh( filename_new, int(cfg.get('basic','срок годности')))):
            return False
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            log.info('----------------------- Processing '+cfgFName )
            cfg = config_read(cfgFName)
            rc_download = False
            if cfg.has_section('download'):
                rc_download = download(cfg)
            filename_in = cfg.get('basic','filename_in')
            if not (rc_download is True or
                    not cfg.has_option('basic','срок годности') or
                    is_file_fresh(filename_in, int(cfg.get('basic', 'срок годности')))):
                return False

            if filename_in == 'new_Price_AUVIX_dealer_csv.csv':
                convert_csv2csv(cfg)
            elif filename_in == 'new_amx_ddp_msk.xlsx':
                convert_excel2csv(cfg)
            else:
                log.error('Не ожидаемый файл %s', filename_in)

if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
